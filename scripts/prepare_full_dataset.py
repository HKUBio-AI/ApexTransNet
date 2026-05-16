from __future__ import annotations

import argparse
import csv
import json
import random
import shutil
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, Iterable, List

import numpy as np
import openpyxl
from PIL import Image, ImageDraw

ANNOTATORS = ["FXY", "LSY", "SMY", "ZSH"]
REVIEW_EXCLUDED_CASES = {"5-43"}


def polygon_to_mask(json_path: Path) -> np.ndarray:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    w = int(data["imageWidth"])
    h = int(data["imageHeight"])
    canvas = Image.new("L", (w, h), 0)
    draw = ImageDraw.Draw(canvas)
    for shape in data.get("shapes", []):
        if shape.get("label") != "AP":
            continue
        points = shape.get("points", []) or []
        if len(points) < 3:
            continue
        draw.polygon([(float(x), float(y)) for x, y in points], outline=255, fill=255)
    return (np.asarray(canvas, dtype=np.uint8) > 0).astype(np.uint8)


def zero_mask_like_image(image_path: Path) -> np.ndarray:
    with Image.open(image_path) as im:
        w, h = im.size
    return np.zeros((h, w), dtype=np.uint8)


def save_u8(mask: np.ndarray, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    Image.fromarray((mask.astype(np.uint8) * 255)).save(path)


def save_float01(arr: np.ndarray, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    Image.fromarray(np.clip(arr * 255.0, 0, 255).astype(np.uint8)).save(path)


def parse_excel(excel_path: Path) -> List[dict]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    batch_map = {"第一批": 1, "第二批": 2, "第三批": 3, "第四批": 4, "第五批": 5}
    current_batch = None
    rows = []
    for row in ws.iter_rows(values_only=True):
        if row[1] in batch_map:
            current_batch = batch_map[row[1]]
            continue
        if current_batch and isinstance(row[1], int):
            data_no = int(row[1])
            case_id = f"{current_batch}-{data_no}"
            diagnosis = str(row[6]).strip() if row[6] is not None else ""
            label_group = "normal" if diagnosis == "正常" else "ap"
            review_flag = "review_excluded" if case_id in REVIEW_EXCLUDED_CASES else ""
            if review_flag:
                split = "exclude"
            else:
                split = ""
            rows.append({
                "case_id": case_id,
                "batch_id": str(current_batch),
                "data_no": str(data_no),
                "check_no": str(row[2]) if row[2] is not None else "",
                "age": str(row[3]) if row[3] is not None else "",
                "sex": str(row[4]) if row[4] is not None else "",
                "tooth": str(row[5]) if row[5] is not None else "",
                "diagnosis": diagnosis,
                "label_group": label_group,
                "review_flag": review_flag,
                "split": split,
            })
    return rows


def stratified_split(rows: List[dict], seed: int = 20260414) -> None:
    rng = random.Random(seed)
    groups: Dict[tuple, List[dict]] = defaultdict(list)
    for row in rows:
        if row["split"] == "exclude":
            continue
        groups[(row["batch_id"], row["label_group"])].append(row)

    for (_, _), items in groups.items():
        rng.shuffle(items)
        n = len(items)
        n_val = max(1, round(n * 0.15))
        n_test = max(1, round(n * 0.15))
        n_train = n - n_val - n_test
        for row in items[:n_train]:
            row["split"] = "train"
        for row in items[n_train:n_train + n_val]:
            row["split"] = "val"
        for row in items[n_train + n_val:]:
            row["split"] = "test"


def find_annotator_json(label_root: Path, batch: str, annotator: str, case_id: str) -> Path | None:
    folder = label_root / f"Periapical periodontitis-{batch}-{annotator}"
    direct = folder / f"{case_id}.json"
    if direct.exists():
        return direct
    matches = list(folder.rglob(f"{case_id}.json"))
    return matches[0] if matches else None


def build_masks_for_case(row: dict, image_path: Path, label_root: Path, processed_root: Path, cache_root: Path) -> dict:
    case_id = row["case_id"]
    batch = row["batch_id"]
    mask_smy_path = processed_root / "masks_smy" / f"{case_id}.png"
    mask_vote2_path = processed_root / "masks_vote2" / f"{case_id}.png"
    soft_path = cache_root / "soft_targets_vote" / f"{case_id}.png"
    unc_path = cache_root / "uncertainty_vote" / f"{case_id}.png"

    if row["review_flag"]:
        base = zero_mask_like_image(image_path)
        save_u8(base, mask_smy_path)
        save_u8(base, mask_vote2_path)
        save_float01(base.astype(np.float32), soft_path)
        save_float01(np.ones_like(base, dtype=np.float32), unc_path)
        return {"num_annotators": "0", "consensus_type": "review_excluded", "sample_weight": "0.0"}

    if row["label_group"] == "normal":
        base = zero_mask_like_image(image_path)
        save_u8(base, mask_smy_path)
        save_u8(base, mask_vote2_path)
        save_float01(base.astype(np.float32), soft_path)
        save_float01(np.ones_like(base, dtype=np.float32), unc_path)
        return {"num_annotators": "0", "consensus_type": "normal_zero", "sample_weight": "1.0"}

    masks = []
    smy_mask = None
    for annotator in ANNOTATORS:
        jp = find_annotator_json(label_root, batch, annotator, case_id)
        if jp is None:
            continue
        m = polygon_to_mask(jp)
        masks.append(m)
        if annotator == "SMY":
            smy_mask = m
    if not masks:
        raise RuntimeError(f"No annotator json found for AP case {case_id}")
    if smy_mask is None:
        smy_mask = masks[0]

    stack = np.stack(masks, axis=0).astype(np.float32)
    votes = stack.sum(axis=0)
    soft = votes / float(stack.shape[0])
    if stack.shape[0] >= 3:
        hard = (votes >= 2).astype(np.uint8)
    elif stack.shape[0] == 2:
        hard = (votes == 2).astype(np.uint8)
    else:
        hard = stack[0].astype(np.uint8)
    agreement = np.maximum(soft, 1.0 - soft)

    save_u8(smy_mask, mask_smy_path)
    save_u8(hard, mask_vote2_path)
    save_float01(soft, soft_path)
    save_float01(agreement, unc_path)
    consensus = {4: "vote2_of4", 3: "vote2_of3", 2: "intersection2", 1: "single_annotator"}.get(stack.shape[0], "custom")
    sample_weight = 0.8 if stack.shape[0] == 2 else 1.0
    return {"num_annotators": str(stack.shape[0]), "consensus_type": consensus, "sample_weight": str(sample_weight)}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--data-root", type=Path, default=Path("/Users/guanjiahui/Desktop/HKU-CRK完整数据"))
    parser.add_argument("--seed", type=int, default=20260414)
    args = parser.parse_args()

    root = args.data_root
    image_root = root / "Summarized data-Periapical periodontitis-X-ray"
    label_root = root / "Label-Periapical periodontitis"
    excel_path = root / "Data-periapical periodontitis-X-ray.xlsx"
    processed_root = root / "processed"
    model_root = root / "AP_model"
    cache_root = model_root / "cache"
    metadata_dir = processed_root / "metadata"
    model_metadata_dir = model_root / "metadata"

    rows = parse_excel(excel_path)
    stratified_split(rows, seed=args.seed)

    fieldnames = [
        "case_id", "batch_id", "data_no", "check_no", "age", "sex", "tooth", "diagnosis",
        "label_group", "review_flag", "annotator", "has_json_label", "is_zero_mask", "image_path",
        "mask_path", "split", "image_path_resolved", "mask_smy_path", "mask_vote2_path",
        "soft_target_path", "uncertainty_path", "num_annotators", "consensus_type", "sample_weight",
        "tooth_root_map_path",
    ]
    out_rows = []
    for row in rows:
        case_id = row["case_id"]
        batch = row["batch_id"]
        src_image = image_root / f"Data-Periapical periodontitis-X-ray-{batch}" / f"{case_id}.jpg"
        if not src_image.exists():
            raise FileNotFoundError(src_image)
        dst_image = processed_root / "images" / f"{case_id}.jpg"
        dst_image.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src_image, dst_image)

        mask_info = build_masks_for_case(row, dst_image, label_root, processed_root, cache_root)
        mask_path = processed_root / "masks_vote2" / f"{case_id}.png"
        out = dict(row)
        out.update({
            "annotator": "multi",
            "has_json_label": "1" if row["label_group"] == "ap" else "0",
            "is_zero_mask": "1" if row["label_group"] == "normal" else "0",
            "image_path": str(src_image),
            "mask_path": str(mask_path),
            "image_path_resolved": str(dst_image),
            "mask_smy_path": str(processed_root / "masks_smy" / f"{case_id}.png"),
            "mask_vote2_path": str(processed_root / "masks_vote2" / f"{case_id}.png"),
            "soft_target_path": str(cache_root / "soft_targets_vote" / f"{case_id}.png"),
            "uncertainty_path": str(cache_root / "uncertainty_vote" / f"{case_id}.png"),
            "tooth_root_map_path": str(cache_root / "tooth_root_maps" / f"{case_id}.png"),
        })
        out.update(mask_info)
        out_rows.append(out)

    metadata_dir.mkdir(parents=True, exist_ok=True)
    model_metadata_dir.mkdir(parents=True, exist_ok=True)
    for out_csv in [metadata_dir / "samples_with_split.csv", model_metadata_dir / "samples_multirater.csv"]:
        with out_csv.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(out_rows)

    summary = {
        "total": len(out_rows),
        "label_group": Counter(r["label_group"] for r in out_rows),
        "split": Counter(r["split"] for r in out_rows),
        "split_by_label": Counter((r["split"], r["label_group"]) for r in out_rows),
        "consensus_type": Counter(r["consensus_type"] for r in out_rows),
        "review_excluded": [r["case_id"] for r in out_rows if r["review_flag"]],
    }
    print("Prepared full dataset")
    for k, v in summary.items():
        print(k, v)
    print("processed_root", processed_root)
    print("metadata", metadata_dir / "samples_with_split.csv")
    print("model_metadata", model_metadata_dir / "samples_multirater.csv")


if __name__ == "__main__":
    main()
